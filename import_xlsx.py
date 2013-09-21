#!/usr/bin/env python
import sys, os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings")
from ht_scoring.scoring.models import Competition, Round, Competitor, Group
from django.utils.encoding import smart_str
import openpyxl

number=0

def get_number(num):
	if not num:
		global number
		number += 1
		return number
	elif int(num) > number:
		number = int(num)
	return int(num)

def splitter(raw, index):
	if raw:
		return raw.split(' ')[index]
	else:
		return ''
		
def remove_non_ascii(s):
	return "".join(i for i in s if ord(i)<128)
		
def auto(raw):
	if raw:
		# return remove_non_ascii(raw)
		return smart_str(raw)
	else:
		return ''

settings={}
settings['worksheet'] = 'Alpha Entries'
settings['start_line'] = 5
settings['fields'] = {}
settings['fields']['first_name']=('A', lambda raw: splitter(raw, -1))
settings['fields']['last_name']=('A', lambda raw: splitter(raw, 0))
settings['fields']['horse']=('B', auto)
settings['fields']['group']=('C', auto)
settings['fields']['comment']=('K', auto)
settings['fields']['number']=('J', get_number)
settings['classes']=[]
#settings['classes'].append({'name': 'Class 1A', 'op_time': 0, 'fences': 10, 'column': 'D'})
settings['classes'].append({'name': 'Class 1 A&B', 'op_time': 0, 'fences': 10, 'column': 'D'})
settings['classes'].append({'name': 'Class 2', 'op_time': 0, 'fences': 10, 'column': 'E'})
settings['classes'].append({'name': 'Class 3', 'op_time': 0, 'fences': 10, 'column': 'F'})
settings['classes'].append({'name': 'Class 4', 'op_time': 0, 'fences': 10, 'column': 'G'})
settings['classes'].append({'name': 'Class 5', 'op_time': 0, 'fences': 10, 'column': 'H'})
settings['classes'].append({'name': 'Class 6', 'op_time': 0, 'fences': 10, 'column': 'I'})

def print_line(row):
	f=row['fields']
	s = 'number: %3d | name: %-20.20s | horse: %-10.10s | pc: %-10.10s | comment: %-50.50s | classes: ' % (f['number'], f['first_name'] + ' ' + f['last_name'], f['horse'], f['group'], remove_non_ascii(f['comment']))
	for cls in row['classes']:
		s+= '%-10.10s ' % cls
	print s

def read_xlsx(fname, settings):
	wb = openpyxl.load_workbook(fname)
	print 'Worksheet names:',wb.get_sheet_names()
	if settings['worksheet'] == '':
		print 'Using Active Sheet: ',
		ws = wb.get_active_sheet()
	else:
		print 'Using Chosen Sheet: ',
		ws = wb.get_sheet_by_name(name = settings['worksheet'])
	print ws.title
	data=[]
	row = settings['start_line']
	while 1:
		row_data={}
		row_data['fields']={}
		keep_going=False
		for key in settings['fields']:
			setting = settings['fields'][key]
			value = ws.cell(setting[0] + str(row)).value
			if value:
				keep_going=True
			value = settings['fields'][key][1](value)
			row_data['fields'][key]=value
		row_data['classes']=[]
		for cls in settings['classes']:
			value = ws.cell(cls['column'] + str(row)).value
			if value:
				row_data['classes'].append(cls['name'])
		if keep_going:
			data.append(row_data)
			row +=1
		else:
			break
	beginning_end = 20
	if beginning_end > len(data) / 2: beginning_end = len(data) / 2
	print 'Finished reading xlsx file'
	print 'First %d rows:' % beginning_end
	for row in data[:beginning_end]:
		print_line(row)
	print 'Last %d rows:' % beginning_end
	for row in data[-beginning_end:]:
		print_line(row)
	print 'Total Entries: %d' % len(data)
	print '           ----------------'
	return data

def add_classes(comptions):
	comption_obs={}
	for comption in comptions:
		name = comption['name']
		comption_filtered = Competition.objects.filter(name = name)
		if comption_filtered:
			print 'Competition \'%s\' already exists' % name
			comption_obs[name] = comption_filtered[0]
		else:
			comption_obs[name] = Competition(name = name)
			comption_obs[name].optimum_time = comption['op_time']
			comption_obs[name].fences = comption['fences']
			comption_obs[name].save()
	return comption_obs

def add_round(competitor, competition):
	if Round.objects.filter(competitor = competitor, competition = competition):
		print 'Round: %d in %s already exists.' % (competitor.number, competition.name)
		return
	the_round = Round(competitor = competitor, competition = competition)
	the_round.not_competative = False
	the_round.eliminated = False
	the_round.save()

def add_row(row, comptions):
	comptors = Competitor.objects.filter(number = row['fields']['number'])
	if comptors:
		print 'Competitor number %d already exists.' % row['fields']['number']
		comptor = comptors[0]
	else:
		groups = Group.objects.filter(name = row['fields']['group'])
		if groups:
			row['fields']['group'] = groups[0]
		else:
			row['fields']['group'] = Group(name = row['fields']['group'])
			row['fields']['group'].save()
		comptor = Competitor(number = row['fields']['number'])
		del row['fields']['number']
		for key in row['fields']:
			setattr(comptor, key, row['fields'][key])
		comptor.save()
	for comption_name in row['classes']:
		add_round(comptor, comptions[comption_name])
	
def add_data(data, settings):
	print 'Adding classes...'
	comptions = add_classes(settings['classes'])
	print 'Adding competitors...'
	for row in data:
		add_row(row, comptions)
	
	
if __name__ == "__main__":
	useage = 'Useage: ./import_xlsx.py file_to_import.xlsx [execute]'
	print 'Script to import xlsx entries into django.', useage
	if len(sys.argv) != 2 and len(sys.argv) != 3:
		print 'Wrong arguments.'
		sys.exit(2)
	fname = sys.argv[1]
	print 'Reading \'%s\'...' % fname
	data = read_xlsx(fname, settings)
	print 'WARNING: You should back up the old db, then remove it and run manager.py syncdb before adding data.'
	inp = raw_input('Are you sure you would like to add this data to the database? [Y/N]  ').lower()
	if inp == 'y':
		print 'Adding data to the database...'
		add_data(data, settings)
		print 'Finished adding data to the database'
		